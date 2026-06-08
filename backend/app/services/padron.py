"""Service layer for padron de alumnos (roster) operations.

Provides:
- ``import_roster``: Parse CSV/XLSX, detect columns, validate, and atomically
  import into a new version.
- ``get_active_entries``: List active entries for a materia.
- ``get_version_history``: List version history for a materia.

Design (per C-06 design.md):
- D1: File processing in the service layer, not the router.
- D2: Column detection by simple header normalization (lowercase + strip).
- D3: Atomic versioning — create version, deactivate previous, bulk insert.
- D4: ``usuario_id`` nullable in EntradaPadron.
- D5: Permission check in router via ``require_permission``.
- D6: Email encrypted via ``EncryptedString`` (model layer).
- D7: In-memory file processing (no temp files).
"""

import csv
import io
import logging
from typing import Any

from fastapi import HTTPException
from starlette.status import (
    HTTP_400_BAD_REQUEST,
    HTTP_404_NOT_FOUND,
    HTTP_422_UNPROCESSABLE_ENTITY,
)

from app.core.action_codes import AccionAuditoria
from app.core.unit_of_work import UnitOfWork
from app.schemas.padron import (
    EntradaPadronOut,
    ImportResultOut,
    VersionHistoryOut,
)
from app.services.audit_service import AuditService

logger = logging.getLogger(__name__)

# ── Expected columns ──────────────────────────────────────────────────────────

_REQUIRED_COLUMNS: set[str] = {"nombre", "email"}
_OPTIONAL_COLUMNS: set[str] = {"apellidos", "apellido", "comision", "regional"}
_KNOWN_COLUMNS: set[str] = _REQUIRED_COLUMNS | _OPTIONAL_COLUMNS

# ── Header normalization ─────────────────────────────────────────────────────


def _normalize_header(header: str) -> str:
    """Normalize a column header: lowercase, strip whitespace, remove BOM.

    Args:
        header: Raw header string from file.

    Returns:
        Normalized header string.
    """
    return header.strip().lower()


def _detect_columns(headers: list[str]) -> dict[str, str]:
    """Detect and validate column mapping from file headers.

    Normalizes each header and builds a mapping from canonical column name
    to original header. Validates that all required columns are present.

    Args:
        headers: List of raw header strings from the file.

    Returns:
        A dict mapping canonical column names to normalized header strings.

    Raises:
        HTTPException(400): If required columns are missing.
    """
    normalized = {_normalize_header(h): h for h in headers}

    # Check required columns
    missing = _REQUIRED_COLUMNS - set(normalized.keys())
    if missing:
        raise HTTPException(
            status_code=HTTP_400_BAD_REQUEST,
            detail=(
                f"Columnas requeridas faltantes: {', '.join(sorted(missing))}. "
                f"Se requieren: {', '.join(sorted(_REQUIRED_COLUMNS))}"
            ),
        )

    # Build mapping
    mapping: dict[str, str] = {}
    for header_original in headers:
        canon = _normalize_header(header_original)
        if canon == "apellido" or canon == "apellidos":
            mapping["apellidos"] = canon
        elif canon in _KNOWN_COLUMNS:
            mapping[canon] = canon
        else:
            logger.warning("Columna no reconocida ignorada: '%s'", header_original)

    return mapping


# ── File parsers ─────────────────────────────────────────────────────────────


def _parse_csv(
    content: str, col_mapping: dict[str, str]
) -> list[dict[str, Any]]:
    """Parse CSV content (handles both comma and semicolon delimiters).

    Args:
        content: Raw CSV text content.
        col_mapping: Column mapping from _detect_columns.

    Returns:
        List of dicts with canonical column names as keys.
    """
    # Detect delimiter by checking the first line
    first_line = content.split("\n")[0] if content else ""
    delimiter = ";"
    if first_line.count(";") < first_line.count(","):
        delimiter = ","

    reader = csv.DictReader(io.StringIO(content), delimiter=delimiter)

    # Build reverse mapping: original header -> canonical name
    reverse_map: dict[str, str] = {}
    for canon, norm in col_mapping.items():
        # Find the original header that normalized to this
        for raw_h in reader.fieldnames or []:
            if _normalize_header(raw_h) == norm:
                reverse_map[raw_h] = canon
                break

    rows: list[dict[str, Any]] = []
    for row in reader:
        entry: dict[str, Any] = {}
        for raw_h, value in row.items():
            canon = reverse_map.get(raw_h)
            if canon:
                entry[canon] = value.strip() if value else ""
        rows.append(entry)

    return rows


def _parse_xlsx(
    content: bytes, col_mapping: dict[str, str]
) -> list[dict[str, Any]]:
    """Parse XLSX content using openpyxl.

    Args:
        content: Raw XLSX file bytes.
        col_mapping: Column mapping from _detect_columns.

    Returns:
        List of dicts with canonical column names as keys.

    Raises:
        HTTPException(400): If the file cannot be parsed.
    """
    try:
        import openpyxl
    except ImportError:
        raise HTTPException(
            status_code=HTTP_422_UNPROCESSABLE_ENTITY,
            detail="XLSX parsing is not available (openpyxl not installed)",
        )

    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
    except Exception as e:
        raise HTTPException(
            status_code=HTTP_400_BAD_REQUEST,
            detail=f"Error al leer archivo XLSX: {e}",
        )

    ws = wb.active
    if ws is None:
        raise HTTPException(
            status_code=HTTP_400_BAD_REQUEST,
            detail="El archivo XLSX no contiene hojas de calculo",
        )

    rows_iter = ws.iter_rows(values_only=True)

    # Read header row
    try:
        raw_headers = [str(cell) if cell is not None else "" for cell in next(rows_iter)]
    except StopIteration:
        raise HTTPException(
            status_code=HTTP_400_BAD_REQUEST,
            detail="El archivo XLSX esta vacio",
        )

    # Build reverse mapping
    reverse_map: dict[str, str] = {}
    for canon, norm in col_mapping.items():
        for raw_h in raw_headers:
            if _normalize_header(str(raw_h)) == norm:
                reverse_map[raw_h] = canon
                break

    rows: list[dict[str, Any]] = []
    for row in rows_iter:
        entry: dict[str, Any] = {}
        for i, value in enumerate(row):
            if i < len(raw_headers):
                raw_h = raw_headers[i]
                canon = reverse_map.get(raw_h)
                if canon:
                    entry[canon] = str(value).strip() if value is not None else ""
        if entry.get("nombre") or entry.get("email"):
            rows.append(entry)

    return rows


# ── PadronService ─────────────────────────────────────────────────────────────


class PadronService:
    """Business logic for padron de alumnos operations.

    Args:
        uow: The ``UnitOfWork`` instance for data access.
    """

    def __init__(self, uow: UnitOfWork):
        self.uow = uow
        self.tenant_id = uow._tenant_id or ""
        self.repo = uow.padron  # backward compat alias

    # ── Import roster ─────────────────────────────────────────────────

    async def import_roster(
        self,
        materia_id: str,
        cohorte_id: str,
        file_bytes: bytes,
        filename: str,
        usuario_id: str,
    ) -> ImportResultOut:
        """Import a roster file (CSV or XLSX) atomically.

        Steps:
        1. Detect file format from extension.
        2. Parse headers and detect columns.
        3. Parse data rows.
        4. Validate non-empty.
        5. Create new version -> deactivate previous -> bulk insert -> commit.
        6. Log audit entry.

        Args:
            materia_id: The materia UUID.
            cohorte_id: The cohorte UUID.
            file_bytes: Raw file content bytes.
            filename: Original filename (used for format detection).
            usuario_id: The importing user's UUID.

        Returns:
            An ``ImportResultOut`` with version_id and count.

        Raises:
            HTTPException(400): If the file is empty, has missing columns,
                or unsupported format.
        """
        # Validate non-empty
        if not file_bytes:
            raise HTTPException(
                status_code=HTTP_400_BAD_REQUEST,
                detail="El archivo esta vacio",
            )

        # Detect format
        lower_name = filename.lower()
        if lower_name.endswith(".csv"):
            rows = self._import_csv(file_bytes)
        elif lower_name.endswith(".xlsx") or lower_name.endswith(".xls"):
            rows = self._import_xlsx(file_bytes)
        else:
            raise HTTPException(
                status_code=HTTP_400_BAD_REQUEST,
                detail=(
                    f"Formato no soportado: '{filename}'. "
                    f"Use archivos .csv o .xlsx"
                ),
            )

        # Validate non-empty after parsing
        if not rows:
            raise HTTPException(
                status_code=HTTP_400_BAD_REQUEST,
                detail="No se encontraron filas validas en el archivo",
            )

        # Atomic import: create version -> deactivate previous -> bulk insert
        # Deactivate previous active version for this (materia, cohorte)
        previous = await self.repo.get_active_version(
            materia_id, cohorte_id
        )
        if previous is not None:
            await self.repo.deactivate_version(previous.id)

        # Create new active version
        version = await self.repo.create_version({
            "materia_id": materia_id,
            "cohorte_id": cohorte_id,
            "cargado_por": usuario_id,
        })

        # Bulk insert entries
        created = await self.repo.create_entries(version.id, rows)

        # Log audit (outside transaction, audit is append-only)
        audit_service = AuditService(self.uow)
        await audit_service.log_action(
            accion=AccionAuditoria.PADRON_CARGAR,
            actor_id=usuario_id,
            tenant_id=self.tenant_id,
            materia_id=materia_id,
            filas_afectadas=created,
        )

        return ImportResultOut(
            version_id=version.id,
            total_imported=created,
        )

    def _import_csv(self, file_bytes: bytes) -> list[dict[str, Any]]:
        """Parse CSV content and return list of entry dicts.

        Args:
            file_bytes: Raw file bytes.

        Returns:
            List of entry dicts with canonical column names.

        Raises:
            HTTPException(400): If parsing fails.
        """
        try:
            content = file_bytes.decode("utf-8-sig")
        except UnicodeDecodeError:
            try:
                content = file_bytes.decode("latin-1")
            except UnicodeDecodeError as e:
                raise HTTPException(
                    status_code=HTTP_400_BAD_REQUEST,
                    detail=f"Error de codificacion: {e}",
                )

        # Detect columns from headers
        first_line = content.split("\n")[0] if content else ""
        if not first_line.strip():
            raise HTTPException(
                status_code=HTTP_400_BAD_REQUEST,
                detail="El archivo CSV esta vacio",
            )

        headers = [h.strip() for h in first_line.split(",")]
        # Also try semicolon
        if len(headers) < 2:
            headers = [h.strip() for h in first_line.split(";")]

        col_mapping = _detect_columns(headers)
        return _parse_csv(content, col_mapping)

    def _import_xlsx(self, file_bytes: bytes) -> list[dict[str, Any]]:
        """Parse XLSX content and return list of entry dicts.

        Args:
            file_bytes: Raw file bytes.

        Returns:
            List of entry dicts with canonical column names.
        """
        import openpyxl

        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True)
        ws = wb.active
        if ws is None:
            raise HTTPException(
                status_code=HTTP_400_BAD_REQUEST,
                detail="El archivo XLSX no contiene hojas de calculo",
            )

        rows_iter = ws.iter_rows(values_only=True)

        try:
            raw_headers = [
                str(cell) if cell is not None else ""
                for cell in next(rows_iter)
            ]
        except StopIteration:
            raise HTTPException(
                status_code=HTTP_400_BAD_REQUEST,
                detail="El archivo XLSX esta vacio",
            )

        col_mapping = _detect_columns(raw_headers)
        return _parse_xlsx(file_bytes, col_mapping)

    # ── Query operations ─────────────────────────────────────────────

    async def get_active_entries(
        self, materia_id: str
    ) -> list[EntradaPadronOut]:
        """Get active entries for a materia.

        Args:
            materia_id: The materia UUID.

        Returns:
            A list of ``EntradaPadronOut`` instances.
        """
        entries = await self.repo.get_active_entries(materia_id)
        return [EntradaPadronOut.model_validate(e) for e in entries]

    async def get_version_history(
        self, materia_id: str
    ) -> list[VersionHistoryOut]:
        """Get version history for a materia.

        Args:
            materia_id: The materia UUID.

        Returns:
            A list of ``VersionHistoryOut`` instances ordered by created_at desc.
        """
        versions = await self.repo.get_version_history(materia_id)
        return [VersionHistoryOut.model_validate(v) for v in versions]
