"""Service layer for calificaciones (grades) operations.

Provides:
- ``preview_import``: Parse file, detect activities, return metadata (no persist).
- ``confirm_import``: Parse file, filter by selected activities, persist atomically.
- ``list_calificaciones``: List grades with optional filters.
- ``clear_calificaciones``: Scope-isolated delete.
- ``set_threshold`` / ``get_threshold``: UmbralMateria CRUD with defaults.

Design (per C-07 design.md):
- D1: Two-phase import via same endpoint (preview=true/false).
- D2: Columns ending in ``(Real)`` detected as numeric grades.
- D3: Threshold defaults: 60%, ["Satisfactorio", "Supera lo esperado"].
- D4: ``aprobado`` derived at import time in the service layer.
- D5: Scope-isolated DELETE by (usuario_id x materia_id).
- D8: File parsing reuses C-06 pattern (openpyxl / csv stdlib).
"""

import csv
import io
import logging
from typing import Any

from fastapi import HTTPException
from starlette.status import HTTP_400_BAD_REQUEST, HTTP_404_NOT_FOUND

from app.core.action_codes import AccionAuditoria
from app.core.unit_of_work import UnitOfWork
from app.schemas.calificaciones import (
    CalificacionListResponse,
    CalificacionResponse,
    ImportResultOut,
    PreviewActividadOut,
    PreviewResultOut,
    UmbralMateriaOut,
)
from app.services.audit_service import AuditService

logger = logging.getLogger(__name__)

# ── Default threshold values ──────────────────────────────────────────────────

_DEFAULT_UMBRAL_PCT = 60
_DEFAULT_VALORES_APROBATORIOS = ["Satisfactorio", "Supera lo esperado"]


# ── Header normalization ─────────────────────────────────────────────────────


def _normalize_header(header: str) -> str:
    """Normalize a column header: lowercase, strip whitespace.

    Args:
        header: Raw header string from file.

    Returns:
        Normalized header string.
    """
    return header.strip().lower()


# ── Column detection (grade-specific) ────────────────────────────────────────

_COL_REQUIRED = {"nombre", "apellidos", "apellido", "email"}


def _detect_nota_columns(
    headers: list[str],
) -> tuple[list[dict[str, Any]], list[str]]:
    """Detect grade columns in file headers.

    Columns ending in ``(Real)`` (case-insensitive) are detected as numeric
    grade columns. All other non-identifying columns are treated as textual
    grade columns.

    Args:
        headers: List of raw header strings from the file.

    Returns:
        A tuple of ``(actividades, identifying)`` where:
        - ``actividades``: list of dicts with keys ``nombre``, ``tipo``,
          ``columna``, ``header_norm``.
        - ``identifying``: list of normalized headers that identify the
          student (nombre, apellidos, email).

    Raises:
        HTTPException(400): If no grade columns are detected.
    """
    normalized = [_normalize_header(h) for h in headers]

    actividades: list[dict[str, Any]] = []
    identifying: list[str] = []

    for i, header_norm in enumerate(normalized):
        raw_header = headers[i]

        # Skip identifying columns
        if header_norm in _COL_REQUIRED:
            identifying.append(header_norm)
            continue

        # Skip empty columns
        if not header_norm:
            continue

        # Detect if this is a numeric grade column (ends with "(real)")
        if header_norm.endswith("(real)"):
            activity_name = raw_header[: -len("(Real)")].strip()
            # Check if we already have an entry for this activity name
            existing = next(
                (a for a in actividades if a["nombre"] == activity_name),
                None,
            )
            if existing:
                # Numeric takes priority over textual for same name
                existing["tipo"] = "numerica"
                existing["columna"] = raw_header
            else:
                actividades.append({
                    "nombre": activity_name,
                    "tipo": "numerica",
                    "columna": raw_header,
                    "header_norm": header_norm,
                })
        else:
            # Textual grade column
            activity_name = raw_header.strip()
            existing = next(
                (a for a in actividades if a["nombre"] == activity_name),
                None,
            )
            if not existing:
                actividades.append({
                    "nombre": activity_name,
                    "tipo": "textual",
                    "columna": raw_header,
                    "header_norm": header_norm,
                })

    if not actividades:
        raise HTTPException(
            status_code=HTTP_400_BAD_REQUEST,
            detail="No se detectaron columnas de calificaciones en el archivo. "
            "Las columnas de nota deben terminar en '(Real)' para notas "
            "numericas o ser nombres de actividad para notas textuales.",
        )

    # Clean up internal keys before returning
    for act in actividades:
        act.pop("header_norm", None)

    return actividades, identifying


# ── File parsers ─────────────────────────────────────────────────────────────


def _parse_csv_rows(
    content: str, identificando: list[str]
) -> tuple[list[dict[str, Any]], list[str]]:
    """Parse CSV content and extract identifying columns + grade columns.

    Args:
        content: Raw CSV text content.
        identificando: List of normalized identifying column headers.

    Returns:
        A tuple of ``(rows, raw_headers)`` where each row is a dict with
        ``identifying`` keys and raw grade column values.
    """
    first_line = content.split("\n")[0] if content else ""
    delimiter = ";"
    if first_line.count(";") < first_line.count(","):
        delimiter = ","

    reader = csv.DictReader(io.StringIO(content), delimiter=delimiter)
    raw_headers = reader.fieldnames or []

    rows: list[dict[str, Any]] = []
    for row in reader:
        entry: dict[str, Any] = {}
        for raw_h, value in row.items():
            if raw_h:
                entry[raw_h] = value.strip() if value else ""
        rows.append(entry)

    return rows, raw_headers


def _parse_xlsx_rows(
    content: bytes,
) -> tuple[list[dict[str, Any]], list[str]]:
    """Parse XLSX content and extract rows and headers.

    Args:
        content: Raw XLSX file bytes.

    Returns:
        A tuple of ``(rows, raw_headers)``.

    Raises:
        HTTPException(400): If the file cannot be parsed.
    """
    try:
        import openpyxl
    except ImportError:
        raise HTTPException(
            status_code=HTTP_400_BAD_REQUEST,
            detail="XLSX parsing is not available (openpyxl not installed)",
        )

    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
    except Exception as e:
        raise HTTPException(
            status_code=HTTP_400_BAD_REQUEST,
            detail=f"Error al leer archivo XLSX: {e}",
        )

    ws = wb.active
    if ws is None:
        raise HTTPException(
            status_code=HTTP_400_BAD_REQUEST,
            detail="El archivo XLSX no contiene hojas de calculo",
        )

    rows_iter = ws.iter_rows(values_only=True)

    try:
        raw_headers = [
            str(cell) if cell is not None else ""
            for cell in next(rows_iter)
        ]
    except StopIteration:
        raise HTTPException(
            status_code=HTTP_400_BAD_REQUEST,
            detail="El archivo XLSX esta vacio",
        )

    rows: list[dict[str, Any]] = []
    for row in rows_iter:
        entry: dict[str, Any] = {}
        for i, value in enumerate(row):
            if i < len(raw_headers) and raw_headers[i]:
                entry[raw_headers[i]] = (
                    str(value).strip() if value is not None else ""
                )
        # Only include rows with at least one non-empty value
        if any(v for v in entry.values()):
            rows.append(entry)

    return rows, raw_headers


# ── Grade derivation ──────────────────────────────────────────────────────────


def _derivar_aprobado(
    nota_numerica: float | None,
    nota_textual: str | None,
    umbral_pct: int,
    valores_aprobatorios: list[str],
) -> bool:
    """Derive ``aprobado`` from grade data and threshold config.

    RN-03: If ``nota_numerica`` exists, compare against ``umbral_pct``.
    RN-02: If only ``nota_textual`` exists, check against
    ``valores_aprobatorios``.

    Args:
        nota_numerica: Numeric grade value (or None).
        nota_textual: Textual grade value (or None).
        umbral_pct: Percentage threshold (0-100).
        valores_aprobatorios: List of approved textual values.

    Returns:
        ``True`` if the grade is approved, ``False`` otherwise.
    """
    if nota_numerica is not None:
        return nota_numerica >= umbral_pct

    if nota_textual:
        return nota_textual.strip().lower() in [
            v.strip().lower() for v in valores_aprobatorios
        ]

    return False


# ── CalificacionesService ─────────────────────────────────────────────────────


class CalificacionesService:
    """Business logic for calificaciones operations.

    Args:
        uow: The ``UnitOfWork`` instance for data access.
    """

    def __init__(self, uow: UnitOfWork):
        self.uow = uow
        self.tenant_id = uow._tenant_id or ""
        self.repo = uow.calificaciones  # backward compat alias

    # ── Import: Preview ───────────────────────────────────────────────

    async def preview_import(
        self,
        materia_id: str,
        file_bytes: bytes,
        filename: str,
    ) -> PreviewResultOut:
        """Parse file and return detected activities without persisting.

        Args:
            materia_id: The materia UUID.
            file_bytes: Raw file content bytes.
            filename: Original filename.

        Returns:
            A ``PreviewResultOut`` with detected activities.

        Raises:
            HTTPException(400): If file is empty or format not supported.
        """
        rows, raw_headers = self._parse_file(file_bytes, filename)
        actividades, _ = _detect_nota_columns(raw_headers)

        # Count total alumnos
        total_alumnos = len(rows)

        # Build preview response with per-activity student counts
        preview_actividades = []
        for act in actividades:
            count = 0
            col = act["columna"]
            for row in rows:
                val = row.get(col, "")
                if val:
                    count += 1
            preview_actividades.append(
                PreviewActividadOut(
                    nombre=act["nombre"],
                    tipo=act["tipo"],
                    columna=act["columna"],
                    total_alumnos=count if count > 0 else total_alumnos,
                )
            )

        return PreviewResultOut(
            actividades=preview_actividades,
            total_alumnos=total_alumnos,
            archivo=filename,
        )

    # ── Import: Confirm ───────────────────────────────────────────────

    async def confirm_import(
        self,
        materia_id: str,
        file_bytes: bytes,
        filename: str,
        actividades_seleccionadas: list[str],
        usuario_id: str,
        asignacion_id: str | None = None,
    ) -> ImportResultOut:
        """Parse file, filter by selected activities, persist atomically.

        Steps:
        1. Parse the file and detect columns.
        2. Get the active padron version to map student entries.
        3. For each selected activity, create Calificacion records.
        4. Derive ``aprobado`` for each record based on threshold.
        5. Bulk insert atomically.
        6. Log audit entry.

        Args:
            materia_id: The materia UUID.
            file_bytes: Raw file content bytes.
            filename: Original filename.
            actividades_seleccionadas: List of activity names to import.
            usuario_id: The importing user's UUID.
            asignacion_id: The user's asignacion UUID for this materia
                (used for threshold lookup). If None, defaults apply.

        Returns:
            An ``ImportResultOut`` with total_importados and actividades list.

        Raises:
            HTTPException(400): If file is empty, format not supported, or
                no activities selected.
        """
        if not actividades_seleccionadas:
            raise HTTPException(
                status_code=HTTP_400_BAD_REQUEST,
                detail="Debe seleccionar al menos una actividad",
            )

        rows, raw_headers = self._parse_file(file_bytes, filename)
        if not rows:
            raise HTTPException(
                status_code=HTTP_400_BAD_REQUEST,
                detail="No se encontraron filas con datos en el archivo",
            )

        actividades, identifying = _detect_nota_columns(raw_headers)

        # Validate selected activities exist in the file
        actividad_names = {a["nombre"] for a in actividades}
        selected_set = set(actividades_seleccionadas)
        invalid = selected_set - actividad_names
        if invalid:
            raise HTTPException(
                status_code=HTTP_400_BAD_REQUEST,
                detail=(
                    f"Actividades no encontradas en el archivo: "
                    f"{', '.join(sorted(invalid))}"
                ),
            )

        # Filter to only selected activities
        selected_actividades = [
            a for a in actividades if a["nombre"] in selected_set
        ]

        # Get active padron entries to map students
        active_version = await self.repo.get_active_version_for_materia(
            materia_id
        )
        if active_version is None:
            raise HTTPException(
                status_code=HTTP_400_BAD_REQUEST,
                detail="No hay un padron activo para esta materia. "
                "Debe importar el padron de alumnos primero.",
            )

        # Get entries for the active version
        entradas = await self.uow.padron.get_entries(active_version.id)

        if not entradas:
            raise HTTPException(
                status_code=HTTP_400_BAD_REQUEST,
                detail="El padron activo no contiene alumnos. "
                "Importe el padron primero.",
            )

        # Get threshold
        umbral = await self._obtener_umbral(asignacion_id, materia_id)

        # Build Calificacion entries
        calificaciones_data: list[dict] = []
        for entry in entradas:
            for act in selected_actividades:
                val = None
                nota_numerica = None
                nota_textual = None

                for row in rows:
                    # Match by email (the most reliable identifier)
                    row_email = row.get("email", row.get("Email", "")).strip().lower()
                    entry_email = (entry.email or "").strip().lower()
                    if row_email and entry_email and row_email == entry_email:
                        val = row.get(act["columna"], "")
                        break

                if not val:
                    continue

                if act["tipo"] == "numerica":
                    try:
                        nota_numerica = float(val.replace(",", "."))
                    except (ValueError, AttributeError):
                        nota_textual = str(val)
                else:
                    nota_textual = str(val)

                aprobado = _derivar_aprobado(
                    nota_numerica=nota_numerica,
                    nota_textual=nota_textual,
                    umbral_pct=umbral["umbral_pct"],
                    valores_aprobatorios=umbral["valores_aprobatorios"],
                )

                calificaciones_data.append({
                    "entrada_padron_id": entry.id,
                    "materia_id": materia_id,
                    "actividad": act["nombre"],
                    "nota_numerica": nota_numerica,
                    "nota_textual": nota_textual,
                    "aprobado": aprobado,
                    "cargado_por": usuario_id,
                })

        if not calificaciones_data:
            raise HTTPException(
                status_code=HTTP_400_BAD_REQUEST,
                detail=(
                    "No se pudieron crear calificaciones. "
                    "Verifique que los emails del archivo coincidan "
                    "con los del padron activo."
                ),
            )

        # Bulk insert atomically
        created = await self.repo.bulk_create_calificaciones(
            calificaciones_data
        )

        # Log audit (outside transaction)
        audit_service = AuditService(self.uow)
        await audit_service.log_action(
            accion=AccionAuditoria.CALIFICACIONES_IMPORTAR,
            actor_id=usuario_id,
            tenant_id=self.tenant_id,
            materia_id=materia_id,
            filas_afectadas=created,
        )

        return ImportResultOut(
            total_importados=created,
            actividades=list(selected_set),
        )

    # ── File parsing ──────────────────────────────────────────────────

    def _parse_file(
        self, file_bytes: bytes, filename: str
    ) -> tuple[list[dict[str, Any]], list[str]]:
        """Parse an uploaded file (CSV or XLSX) and return rows and headers.

        Args:
            file_bytes: Raw file content bytes.
            filename: Original filename.

        Returns:
            A tuple of ``(rows, raw_headers)``.

        Raises:
            HTTPException(400): If format is not supported or file is empty.
        """
        if not file_bytes:
            raise HTTPException(
                status_code=HTTP_400_BAD_REQUEST,
                detail="El archivo esta vacio",
            )

        lower_name = filename.lower()
        if lower_name.endswith(".csv"):
            return self._parse_csv(file_bytes)
        elif lower_name.endswith(".xlsx") or lower_name.endswith(".xls"):
            return self._parse_xlsx(file_bytes)
        else:
            raise HTTPException(
                status_code=HTTP_400_BAD_REQUEST,
                detail="Formato de archivo no soportado. Use .xlsx o .csv",
            )

    def _parse_csv(
        self, file_bytes: bytes
    ) -> tuple[list[dict[str, Any]], list[str]]:
        """Parse CSV content.

        Args:
            file_bytes: Raw file bytes.

        Returns:
            A tuple of ``(rows, raw_headers)``.
        """
        try:
            content = file_bytes.decode("utf-8-sig")
        except UnicodeDecodeError:
            try:
                content = file_bytes.decode("latin-1")
            except UnicodeDecodeError as e:
                raise HTTPException(
                    status_code=HTTP_400_BAD_REQUEST,
                    detail=f"Error de codificacion: {e}",
                )

        first_line = content.split("\n")[0] if content else ""
        if not first_line.strip():
            raise HTTPException(
                status_code=HTTP_400_BAD_REQUEST,
                detail="El archivo CSV esta vacio",
            )

        delimiter = ";"
        if first_line.count(";") < first_line.count(","):
            delimiter = ","

        reader = csv.DictReader(io.StringIO(content), delimiter=delimiter)
        raw_headers = reader.fieldnames or []

        rows: list[dict[str, Any]] = []
        for row in reader:
            entry: dict[str, Any] = {}
            for raw_h, value in row.items():
                if raw_h:
                    entry[raw_h] = value.strip() if value else ""
            if any(v for v in entry.values()):
                rows.append(entry)

        return rows, raw_headers

    def _parse_xlsx(
        self, file_bytes: bytes
    ) -> tuple[list[dict[str, Any]], list[str]]:
        """Parse XLSX content.

        Args:
            file_bytes: Raw file bytes.

        Returns:
            A tuple of ``(rows, raw_headers)``.
        """
        import openpyxl

        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True)
        ws = wb.active
        if ws is None:
            raise HTTPException(
                status_code=HTTP_400_BAD_REQUEST,
                detail="El archivo XLSX no contiene hojas de calculo",
            )

        rows_iter = ws.iter_rows(values_only=True)

        try:
            raw_headers = [
                str(cell) if cell is not None else ""
                for cell in next(rows_iter)
            ]
        except StopIteration:
            raise HTTPException(
                status_code=HTTP_400_BAD_REQUEST,
                detail="El archivo XLSX esta vacio",
            )

        rows: list[dict[str, Any]] = []
        for row in rows_iter:
            entry: dict[str, Any] = {}
            for i, value in enumerate(row):
                if i < len(raw_headers) and raw_headers[i]:
                    entry[raw_headers[i]] = (
                        str(value).strip() if value is not None else ""
                    )
            if any(v for v in entry.values()):
                rows.append(entry)

        return rows, raw_headers



    async def _obtener_umbral(
        self,
        asignacion_id: str | None,
        materia_id: str,
    ) -> dict:
        """Get threshold config for an (asignacion, materia), or defaults.

        Args:
            asignacion_id: The asignacion UUID (or None).
            materia_id: The materia UUID.

        Returns:
            A dict with keys ``umbral_pct`` and ``valores_aprobatorios``.
        """
        if asignacion_id:
            umbral = await self.repo.get_threshold(
                asignacion_id, materia_id
            )
            if umbral:
                return {
                    "umbral_pct": umbral.umbral_pct,
                    "valores_aprobatorios": umbral.valores_aprobatorios,
                }

        return {
            "umbral_pct": _DEFAULT_UMBRAL_PCT,
            "valores_aprobatorios": list(_DEFAULT_VALORES_APROBATORIOS),
        }

    # ── List grades ───────────────────────────────────────────────────

    async def list_calificaciones(
        self,
        materia_id: str,
        actividad: str | None = None,
        aprobado: bool | None = None,
    ) -> CalificacionListResponse:
        """List grade records for a materia with optional filters.

        Args:
            materia_id: The materia UUID.
            actividad: Optional activity name filter.
            aprobado: Optional approval status filter.

        Returns:
            A ``CalificacionListResponse`` with items and total count.
        """
        filters: dict = {}
        if actividad is not None:
            filters["actividad"] = actividad
        if aprobado is not None:
            filters["aprobado"] = aprobado

        records = await self.repo.get_calificaciones(materia_id, filters)
        items = [
            CalificacionResponse.model_validate(r) for r in records
        ]
        return CalificacionListResponse(
            items=items,
            total=len(items),
        )

    # ── Clear grades (scope-isolated) ─────────────────────────────────

    async def clear_calificaciones(
        self,
        materia_id: str,
        usuario_id: str,
        es_coordinador: bool = False,
    ) -> int:
        """Delete grade records with scope isolation (RN-04).

        Args:
            materia_id: The materia UUID.
            usuario_id: The requesting user's UUID.
            es_coordinador: If True, deletes all records for the materia.
                If False, only deletes records loaded by this user.

        Returns:
            The number of deleted records.
        """
        cargado_por = None if es_coordinador else usuario_id
        return await self.repo.delete_calificaciones_scope(
            materia_id=materia_id,
            cargado_por=cargado_por,
        )

    # ── Threshold management ──────────────────────────────────────────

    async def set_threshold(
        self,
        materia_id: str,
        asignacion_id: str,
        data: dict,
    ) -> UmbralMateriaOut:
        """Create or update an approval threshold.

        Args:
            materia_id: The materia UUID.
            asignacion_id: The asignacion UUID.
            data: Dict with optional keys ``umbral_pct`` and/or
                ``valores_aprobatorios``.

        Returns:
            An ``UmbralMateriaOut`` with current threshold values.
        """
        threshold_data = {
            "asignacion_id": asignacion_id,
            "materia_id": materia_id,
        }

        if "umbral_pct" in data and data["umbral_pct"] is not None:
            threshold_data["umbral_pct"] = data["umbral_pct"]
        if (
            "valores_aprobatorios" in data
            and data["valores_aprobatorios"] is not None
        ):
            threshold_data["valores_aprobatorios"] = data["valores_aprobatorios"]

        umbral = await self.repo.upsert_threshold(threshold_data)
        return UmbralMateriaOut(
            id=umbral.id,
            asignacion_id=umbral.asignacion_id,
            materia_id=umbral.materia_id,
            umbral_pct=umbral.umbral_pct,
            valores_aprobatorios=umbral.valores_aprobatorios,
        )

    async def get_threshold(
        self,
        materia_id: str,
        asignacion_id: str,
    ) -> UmbralMateriaOut:
        """Get threshold config, or defaults if not configured.

        Args:
            materia_id: The materia UUID.
            asignacion_id: The asignacion UUID.

        Returns:
            An ``UmbralMateriaOut`` with current or default values.
        """
        umbral = await self.repo.get_threshold(asignacion_id, materia_id)
        if umbral:
            return UmbralMateriaOut(
                id=umbral.id,
                asignacion_id=umbral.asignacion_id,
                materia_id=umbral.materia_id,
                umbral_pct=umbral.umbral_pct,
                valores_aprobatorios=umbral.valores_aprobatorios,
            )

        return UmbralMateriaOut(
            id=None,
            asignacion_id=asignacion_id,
            materia_id=materia_id,
            umbral_pct=_DEFAULT_UMBRAL_PCT,
            valores_aprobatorios=list(_DEFAULT_VALORES_APROBATORIOS),
        )

    # ── Helper: resolve asignacion for PROFESOR ───────────────────────

    async def get_asignacion_for_materia(
        self, usuario_id: str, materia_id: str
    ) -> tuple[str | None, bool]:
        """Get user's active asignacion for a materia.

        Returns (asignacion_id, es_coordinador).
        If user is COORDINADOR/ADMIN, returns (None, True).
        If user is PROFESOR with active assignment, returns (id, False).
        If no access, returns (None, False).

        Args:
            usuario_id: The user UUID.
            materia_id: The materia UUID.

        Returns:
            A tuple of ``(asignacion_id, es_coordinador)``.
        """
        asignaciones = await self.uow.asignacion.list_by_filters(
            usuario_id=usuario_id,
            materia_id=materia_id,
            vigente=True,
        )
        if asignaciones:
            return asignaciones[0].id, False
        return None, False
